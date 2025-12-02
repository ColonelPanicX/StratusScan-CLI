#!/usr/bin/env python3

"""
===========================
= AWS RESOURCE SCANNER =
===========================

Title: AWS IAM Policy Information Collection Script
Version: v2.0.0
Date: SEP-22-2025

Description:
This script collects comprehensive IAM policy information from AWS environments including
managed policies, inline policies, permission analysis, risk assessment, and usage patterns. The data
is exported to an Excel spreadsheet with AWS-specific naming convention for security
auditing and compliance reporting.

Collected information includes: Policy Name, Type, ARN, Attachments, Permission Summary,
Risk Level, Wildcard Usage, Statement Analysis, and Usage Patterns.
"""

import sys
import datetime
import json
import re
from pathlib import Path

# Add path to import utils module
try:
    # Try to import directly (if utils.py is in Python path)
    import utils
except ImportError:
    # If import fails, try to find the module relative to this script
    script_dir = Path(__file__).parent.absolute()

    # Check if we're in the scripts directory
    if script_dir.name.lower() == 'scripts':
        # Add the parent directory (StratusScan root) to the path
        sys.path.append(str(script_dir.parent))
    else:
        # Add the current directory to the path
        sys.path.append(str(script_dir))

    # Try import again
    try:
        import utils
    except ImportError:
        print("ERROR: Could not import the utils module. Make sure utils.py is in the StratusScan directory.")
        sys.exit(1)



def print_title():
    """
    Print the script title and account information.

    Returns:
        tuple: (account_id, account_name)
    """
    print("====================================================================")
    print("                   AWS RESOURCE SCANNER                            ")
    print("====================================================================")
    print("AWS IAM POLICY INFORMATION COLLECTION")
    print("====================================================================")
    print("Version: v2.0.0                       Date: SEP-22-2025")
    # Detect partition and set environment name
    partition = utils.detect_partition()
    partition_name = "AWS GovCloud (US)" if partition == 'aws-us-gov' else "AWS Commercial"
    
    print(f"Environment: {partition_name}")
    print("====================================================================")

    # Get account information
    account_id, account_name = utils.get_account_info()
    print(f"Account ID: {account_id}")
    print(f"Account Name: {account_name}")
    print("====================================================================")

    return account_id, account_name

def calculate_days_since_updated(last_updated_date):
    """
    Calculate days since policy was last updated.

    Args:
        last_updated_date: Date object or None

    Returns:
        str: Days since last updated or descriptive string
    """
    if last_updated_date is None:
        return "Unknown"

    try:
        # Remove timezone info if present for calculation
        if last_updated_date.tzinfo is not None:
            last_updated_date = last_updated_date.replace(tzinfo=None)

        days_since = (datetime.datetime.now() - last_updated_date).days
        return str(days_since)
    except Exception:
        return "Unknown"

def analyze_policy_document(policy_doc):
    """
    Analyze a policy document to extract key information.

    Args:
        policy_doc: Policy document as dict

    Returns:
        dict: Analysis results
    """
    analysis = {
        'permission_summary': 'Unknown',
        'resource_scope': 'Unknown',
        'has_wildcard_actions': 'No',
        'has_wildcard_resources': 'No',
        'statement_count': 0,
        'condition_usage': 'No',
        'risk_level': 'Low'
    }

    try:
        if not policy_doc or 'Statement' not in policy_doc:
            return analysis

        statements = policy_doc['Statement']
        if not isinstance(statements, list):
            statements = [statements]

        analysis['statement_count'] = len(statements)

        actions = set()
        resources = set()
        has_conditions = False
        has_wildcard_actions = False
        has_wildcard_resources = False

        for statement in statements:
            # Skip Deny statements for permission summary
            if statement.get('Effect') != 'Allow':
                continue

            # Analyze actions
            stmt_actions = statement.get('Action', [])
            if isinstance(stmt_actions, str):
                stmt_actions = [stmt_actions]
            elif not isinstance(stmt_actions, list):
                stmt_actions = []

            for action in stmt_actions:
                actions.add(action)
                if '*' in action:
                    has_wildcard_actions = True

            # Analyze resources
            stmt_resources = statement.get('Resource', [])
            if isinstance(stmt_resources, str):
                stmt_resources = [stmt_resources]
            elif not isinstance(stmt_resources, list):
                stmt_resources = []

            for resource in stmt_resources:
                resources.add(resource)
                if resource == '*':
                    has_wildcard_resources = True

            # Check for conditions
            if 'Condition' in statement:
                has_conditions = True

        # Create permission summary (top actions)
        action_list = list(actions)[:10]  # Limit to top 10 actions
        if len(action_list) > 5:
            analysis['permission_summary'] = ', '.join(action_list[:5]) + f' (+{len(action_list)-5} more)'
        else:
            analysis['permission_summary'] = ', '.join(action_list) if action_list else 'None'

        # Create resource scope summary
        resource_list = list(resources)
        if '*' in resource_list:
            analysis['resource_scope'] = 'All resources (*)'
        elif len(resource_list) > 3:
            analysis['resource_scope'] = f'{len(resource_list)} specific resources'
        else:
            analysis['resource_scope'] = ', '.join(resource_list[:3]) if resource_list else 'None'

        # Set flags
        analysis['has_wildcard_actions'] = 'Yes' if has_wildcard_actions else 'No'
        analysis['has_wildcard_resources'] = 'Yes' if has_wildcard_resources else 'No'
        analysis['condition_usage'] = 'Yes' if has_conditions else 'No'

        # Calculate risk level
        risk_factors = 0
        if has_wildcard_actions:
            risk_factors += 2
        if has_wildcard_resources:
            risk_factors += 2
        if not has_conditions and (has_wildcard_actions or has_wildcard_resources):
            risk_factors += 1

        if risk_factors >= 4:
            analysis['risk_level'] = 'High'
        elif risk_factors >= 2:
            analysis['risk_level'] = 'Medium'
        else:
            analysis['risk_level'] = 'Low'

    except Exception as e:
        utils.log_warning(f"Error analyzing policy document: {e}")

    return analysis

@utils.aws_error_handler("Getting policy entities", default_return=('Unknown', 'Unknown', 'Unknown', 0))
def get_policy_entities(iam_client, policy_arn):
    """
    Get entities (users, groups, roles) attached to a policy.

    Args:
        iam_client: The boto3 IAM client
        policy_arn: The policy ARN

    Returns:
        tuple: (attached_users, attached_groups, attached_roles, total_count)
    """
    users = []
    groups = []
    roles = []

    # Get policy entities
    paginator = iam_client.get_paginator('list_entities_for_policy')
    for page in paginator.paginate(PolicyArn=policy_arn):
        # Users
        for user in page.get('PolicyUsers', []):
            users.append(user['UserName'])

        # Groups
        for group in page.get('PolicyGroups', []):
            groups.append(group['GroupName'])

        # Roles
        for role in page.get('PolicyRoles', []):
            roles.append(role['RoleName'])

    total_count = len(users) + len(groups) + len(roles)

    return (
        ', '.join(users) if users else 'None',
        ', '.join(groups) if groups else 'None',
        ', '.join(roles) if roles else 'None',
        total_count
    )

@utils.aws_error_handler("Collecting managed policies", default_return=[])
def collect_managed_policies(iam_client, include_aws_managed=False):
    """
    Collect customer managed and optionally AWS managed policies.

    Args:
        iam_client: The boto3 IAM client
        include_aws_managed: Whether to include AWS managed policies

    Returns:
        list: List of policy information dictionaries
    """
    policies_data = []

    # Get customer managed policies
    utils.log_info("Collecting customer managed policies...")
    paginator = iam_client.get_paginator('list_policies')

    # Count total policies first
    total_policies = 0
    for page in paginator.paginate(Scope='Local'):
        total_policies += len(page['Policies'])

    if include_aws_managed:
        for page in paginator.paginate(Scope='AWS'):
            total_policies += len(page['Policies'])

    utils.log_info(f"Found {total_policies} managed policies to process")

    processed = 0

    # Process customer managed policies
    paginator = iam_client.get_paginator('list_policies')
    for page in paginator.paginate(Scope='Local'):
        policies = page['Policies']

        for policy in policies:
            processed += 1
            progress = (processed / total_policies) * 100
            policy_name = policy['PolicyName']

            utils.log_info(f"[{progress:.1f}%] Processing policy {processed}/{total_policies}: {policy_name}")

            policy_info = process_managed_policy(iam_client, policy, 'Customer Managed')
            if policy_info:
                policies_data.append(policy_info)

    # Process AWS managed policies if requested
    if include_aws_managed:
        utils.log_info("Processing AWS managed policies...")
        for page in paginator.paginate(Scope='AWS'):
            policies = page['Policies']

            for policy in policies:
                processed += 1
                progress = (processed / total_policies) * 100
                policy_name = policy['PolicyName']

                utils.log_info(f"[{progress:.1f}%] Processing AWS policy {processed}/{total_policies}: {policy_name}")

                policy_info = process_managed_policy(iam_client, policy, 'AWS Managed')
                if policy_info:
                    policies_data.append(policy_info)

    return policies_data

def process_managed_policy(iam_client, policy, policy_type):
    """
    Process a single managed policy.

    Args:
        iam_client: The boto3 IAM client
        policy: Policy metadata
        policy_type: 'Customer Managed' or 'AWS Managed'

    Returns:
        dict: Policy information or None if error
    """
    try:
        policy_name = policy['PolicyName']
        policy_arn = policy['Arn']
        policy_id = policy['PolicyId']
        creation_date = policy['CreateDate'].strftime('%Y-%m-%d %H:%M:%S UTC') if policy['CreateDate'] else "Unknown"
        update_date = policy['UpdateDate'].strftime('%Y-%m-%d %H:%M:%S UTC') if policy['UpdateDate'] else "Unknown"
        days_since_updated = calculate_days_since_updated(policy['UpdateDate'])
        path = policy.get('Path', '/')
        description = policy.get('Description', 'None')
        default_version_id = policy.get('DefaultVersionId', 'v1')

        # Get policy document
        try:
            policy_version = iam_client.get_policy_version(
                PolicyArn=policy_arn,
                VersionId=default_version_id
            )
            policy_doc = policy_version['PolicyVersion']['Document']
            analysis = analyze_policy_document(policy_doc)
        except Exception as e:
            utils.log_warning(f"Could not get policy document for {policy_name}: {e}")
            analysis = {
                'permission_summary': 'Unknown',
                'resource_scope': 'Unknown',
                'has_wildcard_actions': 'Unknown',
                'has_wildcard_resources': 'Unknown',
                'statement_count': 0,
                'condition_usage': 'Unknown',
                'risk_level': 'Unknown'
            }

        # Get attached entities
        attached_users, attached_groups, attached_roles, total_attachments = get_policy_entities(iam_client, policy_arn)

        # Determine usage status
        usage_status = 'Used' if total_attachments > 0 else 'Unused'

        return {
            'Policy Name': policy_name,
            'Policy Type': policy_type,
            'Policy ARN': policy_arn,
            'Policy ID': policy_id,
            'Attached To Count': total_attachments,
            'Attached Users': attached_users,
            'Attached Groups': attached_groups,
            'Attached Roles': attached_roles,
            'Permission Summary': analysis['permission_summary'],
            'Resource Scope': analysis['resource_scope'],
            'Has Wildcard Actions': analysis['has_wildcard_actions'],
            'Has Wildcard Resources': analysis['has_wildcard_resources'],
            'Statement Count': analysis['statement_count'],
            'Condition Usage': analysis['condition_usage'],
            'Version': default_version_id,
            'Default Version ID': default_version_id,
            'Creation Date': creation_date,
            'Last Updated': update_date,
            'Days Since Last Updated': days_since_updated,
            'Path': path,
            'Description': description,
            'Usage Status': usage_status,
            'Risk Level': analysis['risk_level']
        }

    except Exception as e:
        utils.log_warning(f"Error processing policy {policy.get('PolicyName', 'Unknown')}: {e}")
        return None

@utils.aws_error_handler("Collecting inline policies", default_return=[])
def collect_inline_policies(iam_client):
    """
    Collect inline policies from users, groups, and roles.

    Args:
        iam_client: The boto3 IAM client

    Returns:
        list: List of inline policy information dictionaries
    """
    inline_policies = []

    utils.log_info("Collecting inline policies from users, groups, and roles...")

    # Count total entities first
    total_entities = 0
    user_paginator = iam_client.get_paginator('list_users')
    for page in user_paginator.paginate():
        total_entities += len(page['Users'])

    group_paginator = iam_client.get_paginator('list_groups')
    for page in group_paginator.paginate():
        total_entities += len(page['Groups'])

    role_paginator = iam_client.get_paginator('list_roles')
    for page in role_paginator.paginate():
        total_entities += len(page['Roles'])

    utils.log_info(f"Scanning {total_entities} entities for inline policies...")

    processed = 0

    # Process users
    user_paginator = iam_client.get_paginator('list_users')
    for page in user_paginator.paginate():
        for user in page['Users']:
            processed += 1
            progress = (processed / total_entities) * 100
            username = user['UserName']

            utils.log_info(f"[{progress:.1f}%] Checking user {processed}/{total_entities}: {username}")

            try:
                user_policies = iam_client.list_user_policies(UserName=username)
                for policy_name in user_policies['PolicyNames']:
                    policy_info = process_inline_policy(iam_client, 'User', username, policy_name)
                    if policy_info:
                        inline_policies.append(policy_info)
            except Exception as e:
                utils.log_warning(f"Error getting inline policies for user {username}: {e}")

    # Process groups
    group_paginator = iam_client.get_paginator('list_groups')
    for page in group_paginator.paginate():
        for group in page['Groups']:
            processed += 1
            progress = (processed / total_entities) * 100
            groupname = group['GroupName']

            utils.log_info(f"[{progress:.1f}%] Checking group {processed}/{total_entities}: {groupname}")

            try:
                group_policies = iam_client.list_group_policies(GroupName=groupname)
                for policy_name in group_policies['PolicyNames']:
                    policy_info = process_inline_policy(iam_client, 'Group', groupname, policy_name)
                    if policy_info:
                        inline_policies.append(policy_info)
            except Exception as e:
                utils.log_warning(f"Error getting inline policies for group {groupname}: {e}")

    # Process roles
    role_paginator = iam_client.get_paginator('list_roles')
    for page in role_paginator.paginate():
        for role in page['Roles']:
            processed += 1
            progress = (processed / total_entities) * 100
            rolename = role['RoleName']

            utils.log_info(f"[{progress:.1f}%] Checking role {processed}/{total_entities}: {rolename}")

            try:
                role_policies = iam_client.list_role_policies(RoleName=rolename)
                for policy_name in role_policies['PolicyNames']:
                    policy_info = process_inline_policy(iam_client, 'Role', rolename, policy_name)
                    if policy_info:
                        inline_policies.append(policy_info)
            except Exception as e:
                utils.log_warning(f"Error getting inline policies for role {rolename}: {e}")

    return inline_policies

def process_inline_policy(iam_client, entity_type, entity_name, policy_name):
    """
    Process a single inline policy.

    Args:
        iam_client: The boto3 IAM client
        entity_type: 'User', 'Group', or 'Role'
        entity_name: Name of the entity
        policy_name: Name of the inline policy

    Returns:
        dict: Policy information or None if error
    """
    try:
        # Get policy document based on entity type
        if entity_type == 'User':
            response = iam_client.get_user_policy(UserName=entity_name, PolicyName=policy_name)
        elif entity_type == 'Group':
            response = iam_client.get_group_policy(GroupName=entity_name, PolicyName=policy_name)
        elif entity_type == 'Role':
            response = iam_client.get_role_policy(RoleName=entity_name, PolicyName=policy_name)
        else:
            return None

        policy_doc = response['PolicyDocument']
        analysis = analyze_policy_document(policy_doc)

        # Create pseudo-ARN for inline policy
        pseudo_arn = f"inline-policy::{entity_type.lower()}::{entity_name}::{policy_name}"

        # Set attachment info based on entity type
        attached_users = entity_name if entity_type == 'User' else 'None'
        attached_groups = entity_name if entity_type == 'Group' else 'None'
        attached_roles = entity_name if entity_type == 'Role' else 'None'

        return {
            'Policy Name': f"{policy_name} ({entity_type}: {entity_name})",
            'Policy Type': 'Inline',
            'Policy ARN': pseudo_arn,
            'Policy ID': f"inline-{entity_type.lower()}-{entity_name}-{policy_name}",
            'Attached To Count': 1,
            'Attached Users': attached_users,
            'Attached Groups': attached_groups,
            'Attached Roles': attached_roles,
            'Permission Summary': analysis['permission_summary'],
            'Resource Scope': analysis['resource_scope'],
            'Has Wildcard Actions': analysis['has_wildcard_actions'],
            'Has Wildcard Resources': analysis['has_wildcard_resources'],
            'Statement Count': analysis['statement_count'],
            'Condition Usage': analysis['condition_usage'],
            'Version': 'N/A',
            'Default Version ID': 'N/A',
            'Creation Date': 'Unknown',
            'Last Updated': 'Unknown',
            'Days Since Last Updated': 'Unknown',
            'Path': 'N/A',
            'Description': f"Inline policy attached to {entity_type.lower()}: {entity_name}",
            'Usage Status': 'Used',
            'Risk Level': analysis['risk_level']
        }

    except Exception as e:
        utils.log_warning(f"Error processing inline policy {policy_name} for {entity_type} {entity_name}: {e}")
        return None

def export_to_excel(managed_policies, inline_policies, account_id, account_name):
    """
    Export IAM policy data to Excel file with AWS naming convention.

    Args:
        managed_policies: List of managed policy dictionaries
        inline_policies: List of inline policy dictionaries
        account_id: AWS account ID
        account_name: AWS account name

    Returns:
        str: Filename of exported file or None if failed
    """
    if not managed_policies and not inline_policies:
        utils.log_warning("No IAM policy data to export.")
        return None

    try:
        # Import pandas after dependency check
        import pandas as pd

        # Generate filename with AWS identifier
        current_date = datetime.datetime.now().strftime("%m.%d.%Y")

        # Use utils module to generate filename and save data with AWS identifier
        filename = utils.create_export_filename(
            account_name,
            "iam-policies",
            "",
            current_date
        )

        # Create data frames for multi-sheet export
        data_frames = {}

        if managed_policies:
            managed_df = pd.DataFrame(managed_policies)
            managed_df = utils.sanitize_for_export(utils.prepare_dataframe_for_export(managed_df))
            data_frames['Customer Managed Policies'] = managed_df

        if inline_policies:
            inline_df = pd.DataFrame(inline_policies)
            inline_df = utils.sanitize_for_export(utils.prepare_dataframe_for_export(inline_df))
            data_frames['Inline Policies'] = inline_df

        # Create summary data
        all_policies = managed_policies + inline_policies
        all_df = pd.DataFrame(all_policies) if all_policies else pd.DataFrame()

        if not all_df.empty:
            # Calculate summary metrics
            total_policies = len(all_df)
            unused_policies = len(all_df[all_df['Usage Status'] == 'Unused'])
            high_risk_policies = len(all_df[all_df['Risk Level'] == 'High'])
            medium_risk_policies = len(all_df[all_df['Risk Level'] == 'Medium'])
            policies_with_wildcards = len(all_df[(all_df['Has Wildcard Actions'] == 'Yes') | (all_df['Has Wildcard Resources'] == 'Yes')])
            policies_without_conditions = len(all_df[all_df['Condition Usage'] == 'No'])

            # Most attached policy
            max_attachments = all_df['Attached To Count'].max() if 'Attached To Count' in all_df.columns else 0
            most_attached = all_df[all_df['Attached To Count'] == max_attachments]['Policy Name'].iloc[0] if max_attachments > 0 else 'None'

            # Recently created policies (last 30 days) - only for managed policies
            recently_created = 0
            old_policies = 0
            if managed_policies:
                managed_df = pd.DataFrame(managed_policies)
                for _, policy in managed_df.iterrows():
                    try:
                        if policy['Days Since Last Updated'] != 'Unknown':
                            days = int(policy['Days Since Last Updated'])
                            if days <= 30:
                                recently_created += 1
                            elif days > 365:
                                old_policies += 1
                    except (ValueError, TypeError):
                        continue

            summary_data = {
                'Metric': [
                    'Total Policies',
                    'Customer Managed Policies',
                    'Inline Policies',
                    'Unused Policies',
                    'High Risk Policies',
                    'Medium Risk Policies',
                    'Policies with Wildcards',
                    'Policies without Conditions',
                    'Most Attached Policy',
                    'Max Attachments',
                    'Recently Created (30 days)',
                    'Due for Review (>365 days)'
                ],
                'Count/Value': [
                    total_policies,
                    len(managed_policies),
                    len(inline_policies),
                    unused_policies,
                    high_risk_policies,
                    medium_risk_policies,
                    policies_with_wildcards,
                    policies_without_conditions,
                    most_attached,
                    max_attachments,
                    recently_created,
                    old_policies
                ]
            }
        else:
            summary_data = {
                'Metric': ['Total Policies'],
                'Count/Value': [0]
            }

        summary_df = pd.DataFrame(summary_data)
        summary_df = utils.sanitize_for_export(utils.prepare_dataframe_for_export(summary_df))
        data_frames['Summary'] = summary_df

        # Save using utils function for multi-sheet Excel
        output_path = utils.save_multiple_dataframes_to_excel(data_frames, filename)

        if output_path:
            # Add conditional formatting for risk levels
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import PatternFill

                wb = load_workbook(output_path)

                # Format managed policies sheet
                if 'Customer Managed Policies' in wb.sheetnames:
                    ws = wb['Customer Managed Policies']
                    risk_col = None

                    # Find Risk Level column
                    for col in range(1, ws.max_column + 1):
                        if ws.cell(row=1, column=col).value == 'Risk Level':
                            risk_col = col
                            break

                    if risk_col:
                        for row in range(2, ws.max_row + 1):
                            risk_value = ws.cell(row=row, column=risk_col).value
                            if risk_value == 'High':
                                ws.cell(row=row, column=risk_col).fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
                            elif risk_value == 'Medium':
                                ws.cell(row=row, column=risk_col).fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
                            elif risk_value == 'Low':
                                ws.cell(row=row, column=risk_col).fill = PatternFill(start_color='E0FFE0', end_color='E0FFE0', fill_type='solid')

                # Format inline policies sheet
                if 'Inline Policies' in wb.sheetnames:
                    ws = wb['Inline Policies']
                    risk_col = None

                    # Find Risk Level column
                    for col in range(1, ws.max_column + 1):
                        if ws.cell(row=1, column=col).value == 'Risk Level':
                            risk_col = col
                            break

                    if risk_col:
                        for row in range(2, ws.max_row + 1):
                            risk_value = ws.cell(row=row, column=risk_col).value
                            if risk_value == 'High':
                                ws.cell(row=row, column=risk_col).fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
                            elif risk_value == 'Medium':
                                ws.cell(row=row, column=risk_col).fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
                            elif risk_value == 'Low':
                                ws.cell(row=row, column=risk_col).fill = PatternFill(start_color='E0FFE0', end_color='E0FFE0', fill_type='solid')

                wb.save(output_path)

            except Exception as e:
                utils.log_warning(f"Could not apply conditional formatting: {e}")

            utils.log_success("AWS IAM policy data exported successfully!")
            utils.log_info(f"File location: {output_path}")
            utils.log_info(f"Export contains {len(managed_policies)} managed and {len(inline_policies)} inline policies")
            return str(output_path)
        else:
            utils.log_error("Error exporting to Excel. Please check the logs.")
            return None

    except Exception as e:
        utils.log_error("Error exporting to Excel", e)
        return None

def main():
    """
    Main function to orchestrate the IAM policy information collection.
    """
    try:
        # Check dependencies first
        if not utils.ensure_dependencies('pandas', 'openpyxl'):
            return

        # Import pandas after dependency check
        import pandas as pd

        # Print title and get account info
        account_id, account_name = print_title()

        # Ask user about AWS managed policies
        include_aws_managed = input("\nInclude AWS managed policies in the export? (y/n): ").lower().strip() == 'y'

        utils.log_info("Starting IAM policy information collection from AWS...")
        print("====================================================================")

        # Create IAM client
        # iam is a global service - use partition-aware home region

        home_region = utils.get_partition_default_region()

        iam_client = utils.get_boto3_client('iam', region_name=home_region)

        # Collect managed policies
        managed_policies = collect_managed_policies(iam_client, include_aws_managed)

        # Collect inline policies
        inline_policies = collect_inline_policies(iam_client)

        if not managed_policies and not inline_policies:
            utils.log_warning("No IAM policy data collected. Exiting.")
            return

        print("\n====================================================================")
        print("COLLECTION COMPLETE")
        print("====================================================================")

        # Export to Excel
        filename = export_to_excel(managed_policies, inline_policies, account_id, account_name)

        if filename:
            utils.log_info(f"Results exported with AWS compliance markers")
            utils.log_info(f"Total managed policies processed: {len(managed_policies)}")
            utils.log_info(f"Total inline policies processed: {len(inline_policies)}")

            # Display some summary statistics
            all_policies = managed_policies + inline_policies
            if all_policies:
                all_df = pd.DataFrame(all_policies)
                utils.log_info(f"High risk policies: {len(all_df[all_df['Risk Level'] == 'High'])}")
                utils.log_info(f"Policies with wildcard actions: {len(all_df[all_df['Has Wildcard Actions'] == 'Yes'])}")
                utils.log_info(f"Unused policies: {len(all_df[all_df['Usage Status'] == 'Unused'])}")

            print("\nScript execution completed.")
        else:
            utils.log_error("Export failed. Please check the logs.")

    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user.")
        sys.exit(0)
    except Exception as e:
        utils.log_error("Unexpected error occurred", e)
        sys.exit(1)

if __name__ == "__main__":
    main()