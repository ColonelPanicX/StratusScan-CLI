#!/usr/bin/env python3

"""
===========================
= AWS RESOURCE SCANNER =
===========================

Title: AWS IAM Export Script
Date: FEB-22-2026

Description:
This script collects and exports IAM resources from AWS environments. It supports
exporting IAM Users, Roles, Policies, or all resources in a single comprehensive
workbook. All data is exported to Excel format with AWS-specific naming conventions
for security auditing and compliance reporting.

Menu:
  1. IAM Users
  2. IAM Roles
  3. IAM Policies
  4. All IAM Resources (Users + Roles + Policies)
"""

import sys
import datetime
import json
import re
from pathlib import Path
from botocore.exceptions import ClientError

# Add path to import utils module
try:
    import utils
except ImportError:
    script_dir = Path(__file__).parent.absolute()
    if script_dir.name.lower() == 'scripts':
        sys.path.append(str(script_dir.parent))
    else:
        sys.path.append(str(script_dir))
    try:
        import utils
    except ImportError:
        print("ERROR: Could not import the utils module. Make sure utils.py is in the StratusScan directory.")
        sys.exit(1)


# ---------------------------------------------------------------------------
# User helper functions (preserved from iam_export.py)
# ---------------------------------------------------------------------------

def calculate_age_in_days(date_obj):
    """
    Calculate the age of a date object in days.

    Args:
        date_obj: Date object to calculate age for

    Returns:
        int or str: Age in days or descriptive string
    """
    if date_obj is None:
        return "Never"

    if isinstance(date_obj, str):
        return date_obj

    try:
        if date_obj.tzinfo is not None:
            date_obj = date_obj.replace(tzinfo=None)
        age = datetime.datetime.now() - date_obj
        return age.days
    except Exception:
        return "Unknown"


@utils.aws_error_handler("Getting user MFA devices", default_return="Unknown")
def get_user_mfa_devices(iam_client, username):
    """
    Get MFA devices for a user.

    Args:
        iam_client: The boto3 IAM client
        username: The username to check

    Returns:
        str: MFA status (Enabled/Disabled/Unknown)
    """
    virtual_mfa = iam_client.list_mfa_devices(UserName=username)
    mfa_devices = virtual_mfa.get('MFADevices', [])
    return "Enabled" if mfa_devices else "Disabled"


@utils.aws_error_handler("Getting user groups", default_return="Unknown")
def get_user_groups(iam_client, username):
    """
    Get groups that a user belongs to.

    Args:
        iam_client: The boto3 IAM client
        username: The username to check

    Returns:
        str: Comma-separated list of group names or descriptive string
    """
    response = iam_client.get_groups_for_user(UserName=username)
    groups = [group['GroupName'] for group in response['Groups']]
    return ", ".join(groups) if groups else "None"


@utils.aws_error_handler("Getting user policies", default_return="Unknown")
def get_user_policies(iam_client, username):
    """
    Get all policies attached to a user (both attached and inline).

    Args:
        iam_client: The boto3 IAM client
        username: The username to check

    Returns:
        str: Comma-separated list of policy names or descriptive string
    """
    policies = []
    attached_policies = iam_client.list_attached_user_policies(UserName=username)
    for policy in attached_policies['AttachedPolicies']:
        policies.append(policy['PolicyName'])

    inline_policies = iam_client.list_user_policies(UserName=username)
    for policy_name in inline_policies['PolicyNames']:
        policies.append(f"{policy_name} (Inline)")

    return ", ".join(policies) if policies else "None"


def get_password_info(iam_client, username):
    """
    Get password-related information for a user.

    Args:
        iam_client: The boto3 IAM client
        username: The username to check

    Returns:
        tuple: (password_age, console_access)
    """
    try:
        login_profile = iam_client.get_login_profile(UserName=username)
        password_creation = login_profile['LoginProfile']['CreateDate']
        password_age = calculate_age_in_days(password_creation)
        console_access = "Enabled"
        return password_age, console_access
    except ClientError as e:
        if e.response['Error']['Code'] == 'NoSuchEntity':
            return "No Password", "Disabled"
        else:
            return "Unknown", "Unknown"
    except Exception as e:
        utils.log_warning(f"Could not get password info for user {username}: {e}")
        return "Unknown", "Unknown"


@utils.aws_error_handler("Getting access key information", default_return=("Unknown", "Unknown", "Unknown"))
def get_access_key_info(iam_client, username):
    """
    Get access key information for a user.

    Args:
        iam_client: The boto3 IAM client
        username: The username to check

    Returns:
        tuple: (access_key_ids, active_key_age, access_key_last_used)
    """
    response = iam_client.list_access_keys(UserName=username)
    access_keys = response['AccessKeyMetadata']

    if not access_keys:
        return "None", "No Keys", "Never"

    key_info = []
    active_ages = []
    last_used_dates = []

    for key in access_keys:
        key_id = key['AccessKeyId']
        status = key['Status']
        created_date = key['CreateDate']

        key_info.append(f"{key_id} ({status})")

        if status == 'Active':
            key_age = calculate_age_in_days(created_date)
            active_ages.append(str(key_age))

            try:
                last_used_response = iam_client.get_access_key_last_used(AccessKeyId=key_id)
                last_used_date = last_used_response.get('AccessKeyLastUsed', {}).get('LastUsedDate')
                if last_used_date:
                    days_since_used = calculate_age_in_days(last_used_date)
                    last_used_dates.append(f"{days_since_used} days ago")
                else:
                    last_used_dates.append("Never")
            except Exception:
                last_used_dates.append("Unknown")

    access_key_ids = ", ".join(key_info)
    active_key_age = ", ".join(active_ages) if active_ages else "No Active Keys"
    access_key_last_used = ", ".join(last_used_dates) if last_used_dates else "Never"

    return access_key_ids, active_key_age, access_key_last_used


def collect_iam_user_information():
    """
    Collect IAM user information from AWS.

    Returns:
        list: List of dictionaries containing user information
    """
    utils.log_info("Collecting IAM user information from AWS environment...")

    try:
        home_region = utils.get_partition_default_region()
        iam_client = utils.get_boto3_client('iam', region_name=home_region)
    except Exception as e:
        utils.log_error("Error creating IAM client", e)
        return []

    user_data = []

    try:
        paginator = iam_client.get_paginator('list_users')

        total_users = 0
        for page in paginator.paginate():
            total_users += len(page['Users'])

        utils.log_info(f"Found {total_users} IAM users to process")

        paginator = iam_client.get_paginator('list_users')
        processed = 0

        for page in paginator.paginate():
            users = page['Users']

            for user in users:
                username = user['UserName']
                processed += 1
                progress = (processed / total_users) * 100 if total_users > 0 else 0

                utils.log_info(f"[{progress:.1f}%] Processing user {processed}/{total_users}: {username}")

                creation_date = user['CreateDate'].strftime('%Y-%m-%d %H:%M:%S UTC') if user['CreateDate'] else "Unknown"
                password_last_used = user.get('PasswordLastUsed')

                if password_last_used:
                    console_last_signin = password_last_used.strftime('%Y-%m-%d %H:%M:%S UTC')
                else:
                    console_last_signin = "Never"

                groups = get_user_groups(iam_client, username)
                mfa_status = get_user_mfa_devices(iam_client, username)
                password_age, console_access = get_password_info(iam_client, username)
                access_key_id, active_key_age, access_key_last_used = get_access_key_info(iam_client, username)
                permission_policies = get_user_policies(iam_client, username)

                user_info = {
                    'User Name': username,
                    'Groups': groups,
                    'MFA': mfa_status,
                    'Password Age': f"{password_age} days" if isinstance(password_age, int) else password_age,
                    'Console Last Sign-in': console_last_signin,
                    'Access Key ID': access_key_id,
                    'Active Key Age': f"{active_key_age} days" if isinstance(active_key_age, int) else active_key_age,
                    'Access Key Last Used': access_key_last_used,
                    'Creation Date': creation_date,
                    'Console Access': console_access,
                    'Permission Policies': permission_policies
                }

                user_data.append(user_info)

    except Exception as e:
        utils.log_error("Error collecting IAM user information", e)
        return []

    utils.log_success(f"Successfully collected information for {len(user_data)} users")
    return user_data


# ---------------------------------------------------------------------------
# Role helper functions (improved versions from iam_comprehensive_export.py)
# ---------------------------------------------------------------------------

def calculate_days_since_last_used(last_used_date):
    """
    Calculate days since role was last used.

    Args:
        last_used_date: Date object or None

    Returns:
        str: Days since last used or descriptive string
    """
    if last_used_date is None:
        return "Never"

    try:
        if last_used_date.tzinfo is not None:
            last_used_date = last_used_date.replace(tzinfo=None)
        days_since = (datetime.datetime.now() - last_used_date).days
        return str(days_since)
    except Exception:
        return "Unknown"


def analyze_trust_policy(trust_policy_doc):
    """
    Analyze the trust policy to extract key information.
    Handles both arn:aws: and arn:aws-us-gov: partitions.

    Args:
        trust_policy_doc: Trust policy document as dict

    Returns:
        tuple: (trusted_entities, trust_summary, cross_account_info, service_usage)
    """
    trusted_entities = []
    cross_account_accounts = []
    services = []

    try:
        statements = trust_policy_doc.get('Statement', [])
        if not isinstance(statements, list):
            statements = [statements]

        for statement in statements:
            if statement.get('Effect') == 'Allow':
                principals = statement.get('Principal', {})

                if isinstance(principals, str):
                    if principals == '*':
                        trusted_entities.append("Anyone (*)")
                    else:
                        trusted_entities.append(principals)
                elif isinstance(principals, dict):
                    if 'AWS' in principals:
                        aws_principals = principals['AWS']
                        if not isinstance(aws_principals, list):
                            aws_principals = [aws_principals]

                        for principal in aws_principals:
                            if isinstance(principal, str):
                                if principal == '*':
                                    trusted_entities.append("Any AWS Account (*)")
                                elif 'arn:aws' in principal:
                                    # Extract account ID from ARN (handles both partitions)
                                    match = re.search(r':(\d{12}):', principal)
                                    if match:
                                        account_id = match.group(1)
                                        cross_account_accounts.append(account_id)
                                        trusted_entities.append(f"AWS Account: {account_id}")
                                    else:
                                        trusted_entities.append(f"AWS: {principal}")
                                else:
                                    trusted_entities.append(f"AWS: {principal}")

                    if 'Service' in principals:
                        service_principals = principals['Service']
                        if not isinstance(service_principals, list):
                            service_principals = [service_principals]

                        for service in service_principals:
                            services.append(service)
                            trusted_entities.append(f"Service: {service}")

                    if 'Federated' in principals:
                        federated_principals = principals['Federated']
                        if not isinstance(federated_principals, list):
                            federated_principals = [federated_principals]

                        for fed in federated_principals:
                            trusted_entities.append(f"Federated: {fed}")

        trust_summary_parts = []
        if services:
            trust_summary_parts.append(f"Services: {len(services)}")
        if cross_account_accounts:
            trust_summary_parts.append(f"Cross-Account: {len(cross_account_accounts)}")
        if not services and not cross_account_accounts and trusted_entities:
            trust_summary_parts.append("Other principals")

        trust_summary = ", ".join(trust_summary_parts) if trust_summary_parts else "None"

        cross_account_access = "Yes" if cross_account_accounts else "No"
        cross_account_details = ", ".join(cross_account_accounts) if cross_account_accounts else "None"

        return (
            ", ".join(trusted_entities[:5]) + ("..." if len(trusted_entities) > 5 else ""),
            trust_summary,
            f"{cross_account_access} ({cross_account_details})" if cross_account_accounts else "No",
            ", ".join(services) if services else "None"
        )

    except Exception as e:
        utils.log_warning(f"Error analyzing trust policy: {e}")
        return "Unknown", "Unknown", "Unknown", "Unknown"


def determine_role_type(role_name, role_path, trust_policy_doc):
    """
    Determine the type of IAM role.
    Uses utils.get_account_info() for GovCloud-correct cross-account detection.

    Args:
        role_name: Name of the role
        role_path: Path of the role
        trust_policy_doc: Trust policy document

    Returns:
        str: Role type classification
    """
    if role_path.startswith('/aws-service-role/') or 'ServiceLinkedRole' in role_name:
        return "Service-linked"

    try:
        current_account_id, _ = utils.get_account_info()
    except Exception:
        current_account_id = None

    try:
        statements = trust_policy_doc.get('Statement', [])
        if not isinstance(statements, list):
            statements = [statements]

        for statement in statements:
            if statement.get('Effect') == 'Allow':
                principals = statement.get('Principal', {})
                if isinstance(principals, dict) and 'AWS' in principals:
                    aws_principals = principals['AWS']
                    if not isinstance(aws_principals, list):
                        aws_principals = [aws_principals]

                    for principal in aws_principals:
                        # Handle both arn:aws: and arn:aws-us-gov: partitions
                        if isinstance(principal, str) and re.search(r'arn:aws(-us-gov)?:', principal):
                            match = re.search(r':(\d{12}):', principal)
                            if match:
                                principal_account_id = match.group(1)
                                if current_account_id is None or principal_account_id != current_account_id:
                                    return "Cross-account"
    except Exception as e:
        utils.log_warning(f"Could not determine role type from trust policy: {e}")

    return "Standard"


@utils.aws_error_handler("Getting role policies", default_return="Unknown")
def get_role_policies(iam_client, role_name):
    """
    Get all policies attached to a role (both managed and inline).

    Args:
        iam_client: The boto3 IAM client
        role_name: The role name to check

    Returns:
        str: Comma-separated list of policy names
    """
    policies = []

    attached_policies = iam_client.list_attached_role_policies(RoleName=role_name)
    for policy in attached_policies['AttachedPolicies']:
        policies.append(policy['PolicyName'])

    inline_policies = iam_client.list_role_policies(RoleName=role_name)
    for policy_name in inline_policies['PolicyNames']:
        policies.append(f"{policy_name} (Inline)")

    return ", ".join(policies) if policies else "None"


@utils.aws_error_handler("Getting role tags", default_return="Unknown")
def get_role_tags(iam_client, role_name):
    """
    Get tags for a role.

    Args:
        iam_client: The boto3 IAM client
        role_name: The role name to check

    Returns:
        str: Comma-separated list of key=value pairs
    """
    response = iam_client.list_role_tags(RoleName=role_name)
    tags = response.get('Tags', [])
    tag_strings = [f"{tag['Key']}={tag['Value']}" for tag in tags]
    return ", ".join(tag_strings) if tag_strings else "None"


@utils.aws_error_handler("Collecting IAM role information", default_return=[])
def collect_iam_role_information():
    """
    Collect IAM role information from AWS.

    Returns:
        list: List of dictionaries containing role information
    """
    utils.log_info("Collecting IAM role information from AWS environment...")

    home_region = utils.get_partition_default_region()
    iam_client = utils.get_boto3_client('iam', region_name=home_region)
    role_data = []

    paginator = iam_client.get_paginator('list_roles')

    total_roles = 0
    for page in paginator.paginate():
        total_roles += len(page['Roles'])

    utils.log_info(f"Found {total_roles} IAM roles to process")

    paginator = iam_client.get_paginator('list_roles')
    processed = 0

    for page in paginator.paginate():
        roles = page['Roles']

        for role in roles:
            role_name = role['RoleName']
            processed += 1
            progress = (processed / total_roles) * 100 if total_roles > 0 else 0

            utils.log_info(f"[{progress:.1f}%] Processing role {processed}/{total_roles}: {role_name}")

            creation_date = role['CreateDate'].strftime('%Y-%m-%d %H:%M:%S UTC') if role['CreateDate'] else "Unknown"
            role_path = role.get('Path', '/')
            description = role.get('Description', 'None')
            max_session_duration = role.get('MaxSessionDuration', 3600) // 3600

            trust_policy_doc = role.get('AssumeRolePolicyDocument', {})
            trusted_entities, trust_summary, cross_account_info, service_usage = analyze_trust_policy(trust_policy_doc)

            role_type = determine_role_type(role_name, role_path, trust_policy_doc)

            try:
                role_usage = iam_client.get_role(RoleName=role_name)
                role_last_used = role_usage['Role'].get('RoleLastUsed', {})
                last_used_date = role_last_used.get('LastUsedDate')

                if last_used_date:
                    last_used_str = last_used_date.strftime('%Y-%m-%d %H:%M:%S UTC')
                    days_since_used = calculate_days_since_last_used(last_used_date)
                else:
                    last_used_str = "Never"
                    days_since_used = "Never"

            except Exception as e:
                utils.log_warning(f"Could not get usage info for role {role_name}: {e}")
                last_used_str = "Unknown"
                days_since_used = "Unknown"

            permission_policies = get_role_policies(iam_client, role_name)
            tags = get_role_tags(iam_client, role_name)

            role_info = {
                'Role Name': role_name,
                'Role Type': role_type,
                'Trusted Entities': trusted_entities,
                'Trust Policy Summary': trust_summary,
                'Permission Policies': permission_policies,
                'Last Used': last_used_str,
                'Days Since Last Used': days_since_used,
                'Max Session Duration (Hours)': max_session_duration,
                'Cross-Account Access': cross_account_info,
                'Service Usage': service_usage,
                'Creation Date': creation_date,
                'Path': role_path,
                'Description': description,
                'Tags': tags
            }

            role_data.append(role_info)

    utils.log_success(f"Successfully collected information for {len(role_data)} roles")
    return role_data


# ---------------------------------------------------------------------------
# Policy helper functions (preserved from iam_policies_export.py)
# ---------------------------------------------------------------------------

def calculate_days_since_updated(last_updated_date):
    """
    Calculate days since policy was last updated.

    Args:
        last_updated_date: Date object or None

    Returns:
        str: Days since last updated or descriptive string
    """
    if last_updated_date is None:
        return "Unknown"

    try:
        if last_updated_date.tzinfo is not None:
            last_updated_date = last_updated_date.replace(tzinfo=None)
        days_since = (datetime.datetime.now() - last_updated_date).days
        return str(days_since)
    except Exception:
        return "Unknown"


def analyze_policy_document(policy_doc):
    """
    Analyze a policy document to extract key information.

    Args:
        policy_doc: Policy document as dict

    Returns:
        dict: Analysis results
    """
    analysis = {
        'permission_summary': 'Unknown',
        'resource_scope': 'Unknown',
        'has_wildcard_actions': 'No',
        'has_wildcard_resources': 'No',
        'statement_count': 0,
        'condition_usage': 'No',
        'risk_level': 'Low'
    }

    try:
        if not policy_doc or 'Statement' not in policy_doc:
            return analysis

        statements = policy_doc['Statement']
        if not isinstance(statements, list):
            statements = [statements]

        analysis['statement_count'] = len(statements)

        actions = set()
        resources = set()
        has_conditions = False
        has_wildcard_actions = False
        has_wildcard_resources = False

        for statement in statements:
            if statement.get('Effect') != 'Allow':
                continue

            stmt_actions = statement.get('Action', [])
            if isinstance(stmt_actions, str):
                stmt_actions = [stmt_actions]
            elif not isinstance(stmt_actions, list):
                stmt_actions = []

            for action in stmt_actions:
                actions.add(action)
                if '*' in action:
                    has_wildcard_actions = True

            stmt_resources = statement.get('Resource', [])
            if isinstance(stmt_resources, str):
                stmt_resources = [stmt_resources]
            elif not isinstance(stmt_resources, list):
                stmt_resources = []

            for resource in stmt_resources:
                resources.add(resource)
                if resource == '*':
                    has_wildcard_resources = True

            if 'Condition' in statement:
                has_conditions = True

        action_list = list(actions)[:10]
        if len(action_list) > 5:
            analysis['permission_summary'] = ', '.join(action_list[:5]) + f' (+{len(action_list)-5} more)'
        else:
            analysis['permission_summary'] = ', '.join(action_list) if action_list else 'None'

        resource_list = list(resources)
        if '*' in resource_list:
            analysis['resource_scope'] = 'All resources (*)'
        elif len(resource_list) > 3:
            analysis['resource_scope'] = f'{len(resource_list)} specific resources'
        else:
            analysis['resource_scope'] = ', '.join(resource_list[:3]) if resource_list else 'None'

        analysis['has_wildcard_actions'] = 'Yes' if has_wildcard_actions else 'No'
        analysis['has_wildcard_resources'] = 'Yes' if has_wildcard_resources else 'No'
        analysis['condition_usage'] = 'Yes' if has_conditions else 'No'

        risk_factors = 0
        if has_wildcard_actions:
            risk_factors += 2
        if has_wildcard_resources:
            risk_factors += 2
        if not has_conditions and (has_wildcard_actions or has_wildcard_resources):
            risk_factors += 1

        if risk_factors >= 4:
            analysis['risk_level'] = 'High'
        elif risk_factors >= 2:
            analysis['risk_level'] = 'Medium'
        else:
            analysis['risk_level'] = 'Low'

    except Exception as e:
        utils.log_warning(f"Error analyzing policy document: {e}")

    return analysis


@utils.aws_error_handler("Getting policy entities", default_return=('Unknown', 'Unknown', 'Unknown', 0))
def get_policy_entities(iam_client, policy_arn):
    """
    Get entities (users, groups, roles) attached to a policy.

    Args:
        iam_client: The boto3 IAM client
        policy_arn: The policy ARN

    Returns:
        tuple: (attached_users, attached_groups, attached_roles, total_count)
    """
    users = []
    groups = []
    roles = []

    paginator = iam_client.get_paginator('list_entities_for_policy')
    for page in paginator.paginate(PolicyArn=policy_arn):
        for user in page.get('PolicyUsers', []):
            users.append(user['UserName'])
        for group in page.get('PolicyGroups', []):
            groups.append(group['GroupName'])
        for role in page.get('PolicyRoles', []):
            roles.append(role['RoleName'])

    total_count = len(users) + len(groups) + len(roles)

    return (
        ', '.join(users) if users else 'None',
        ', '.join(groups) if groups else 'None',
        ', '.join(roles) if roles else 'None',
        total_count
    )


@utils.aws_error_handler("Collecting managed policies", default_return=[])
def collect_managed_policies(iam_client, include_aws_managed=False):
    """
    Collect customer managed and optionally AWS managed policies.

    Args:
        iam_client: The boto3 IAM client
        include_aws_managed: Whether to include AWS managed policies

    Returns:
        list: List of policy information dictionaries
    """
    policies_data = []

    utils.log_info("Collecting customer managed policies...")
    paginator = iam_client.get_paginator('list_policies')

    total_policies = 0
    for page in paginator.paginate(Scope='Local'):
        total_policies += len(page['Policies'])

    if include_aws_managed:
        for page in paginator.paginate(Scope='AWS'):
            total_policies += len(page['Policies'])

    utils.log_info(f"Found {total_policies} managed policies to process")

    processed = 0

    paginator = iam_client.get_paginator('list_policies')
    for page in paginator.paginate(Scope='Local'):
        policies = page['Policies']
        for policy in policies:
            processed += 1
            progress = (processed / total_policies) * 100 if total_policies > 0 else 0
            policy_name = policy['PolicyName']
            utils.log_info(f"[{progress:.1f}%] Processing policy {processed}/{total_policies}: {policy_name}")
            policy_info = process_managed_policy(iam_client, policy, 'Customer Managed')
            if policy_info:
                policies_data.append(policy_info)

    if include_aws_managed:
        utils.log_info("Processing AWS managed policies...")
        for page in paginator.paginate(Scope='AWS'):
            policies = page['Policies']
            for policy in policies:
                processed += 1
                progress = (processed / total_policies) * 100 if total_policies > 0 else 0
                policy_name = policy['PolicyName']
                utils.log_info(f"[{progress:.1f}%] Processing AWS policy {processed}/{total_policies}: {policy_name}")
                policy_info = process_managed_policy(iam_client, policy, 'AWS Managed')
                if policy_info:
                    policies_data.append(policy_info)

    return policies_data


def process_managed_policy(iam_client, policy, policy_type):
    """
    Process a single managed policy.

    Args:
        iam_client: The boto3 IAM client
        policy: Policy metadata
        policy_type: 'Customer Managed' or 'AWS Managed'

    Returns:
        dict: Policy information or None if error
    """
    try:
        policy_name = policy['PolicyName']
        policy_arn = policy['Arn']
        policy_id = policy['PolicyId']
        creation_date = policy['CreateDate'].strftime('%Y-%m-%d %H:%M:%S UTC') if policy['CreateDate'] else "Unknown"
        update_date = policy['UpdateDate'].strftime('%Y-%m-%d %H:%M:%S UTC') if policy['UpdateDate'] else "Unknown"
        days_since_updated = calculate_days_since_updated(policy['UpdateDate'])
        path = policy.get('Path', '/')
        description = policy.get('Description', 'None')
        default_version_id = policy.get('DefaultVersionId', 'v1')

        try:
            policy_version = iam_client.get_policy_version(
                PolicyArn=policy_arn,
                VersionId=default_version_id
            )
            policy_doc = policy_version['PolicyVersion']['Document']
            analysis = analyze_policy_document(policy_doc)
        except Exception as e:
            utils.log_warning(f"Could not get policy document for {policy_name}: {e}")
            analysis = {
                'permission_summary': 'Unknown',
                'resource_scope': 'Unknown',
                'has_wildcard_actions': 'Unknown',
                'has_wildcard_resources': 'Unknown',
                'statement_count': 0,
                'condition_usage': 'Unknown',
                'risk_level': 'Unknown'
            }

        attached_users, attached_groups, attached_roles, total_attachments = get_policy_entities(iam_client, policy_arn)
        usage_status = 'Used' if total_attachments > 0 else 'Unused'

        return {
            'Policy Name': policy_name,
            'Policy Type': policy_type,
            'Policy ARN': policy_arn,
            'Policy ID': policy_id,
            'Attached To Count': total_attachments,
            'Attached Users': attached_users,
            'Attached Groups': attached_groups,
            'Attached Roles': attached_roles,
            'Permission Summary': analysis['permission_summary'],
            'Resource Scope': analysis['resource_scope'],
            'Has Wildcard Actions': analysis['has_wildcard_actions'],
            'Has Wildcard Resources': analysis['has_wildcard_resources'],
            'Statement Count': analysis['statement_count'],
            'Condition Usage': analysis['condition_usage'],
            'Version': default_version_id,
            'Default Version ID': default_version_id,
            'Creation Date': creation_date,
            'Last Updated': update_date,
            'Days Since Last Updated': days_since_updated,
            'Path': path,
            'Description': description,
            'Usage Status': usage_status,
            'Risk Level': analysis['risk_level']
        }

    except Exception as e:
        utils.log_warning(f"Error processing policy {policy.get('PolicyName', 'Unknown')}: {e}")
        return None


@utils.aws_error_handler("Collecting inline policies", default_return=[])
def collect_inline_policies(iam_client):
    """
    Collect inline policies from users, groups, and roles.

    Args:
        iam_client: The boto3 IAM client

    Returns:
        list: List of inline policy information dictionaries
    """
    inline_policies = []

    utils.log_info("Collecting inline policies from users, groups, and roles...")

    total_entities = 0
    user_paginator = iam_client.get_paginator('list_users')
    for page in user_paginator.paginate():
        total_entities += len(page['Users'])

    group_paginator = iam_client.get_paginator('list_groups')
    for page in group_paginator.paginate():
        total_entities += len(page['Groups'])

    role_paginator = iam_client.get_paginator('list_roles')
    for page in role_paginator.paginate():
        total_entities += len(page['Roles'])

    utils.log_info(f"Scanning {total_entities} entities for inline policies...")

    processed = 0

    user_paginator = iam_client.get_paginator('list_users')
    for page in user_paginator.paginate():
        for user in page['Users']:
            processed += 1
            progress = (processed / total_entities) * 100 if total_entities > 0 else 0
            username = user['UserName']
            utils.log_info(f"[{progress:.1f}%] Checking user {processed}/{total_entities}: {username}")

            try:
                user_policies = iam_client.list_user_policies(UserName=username)
                for policy_name in user_policies['PolicyNames']:
                    policy_info = process_inline_policy(iam_client, 'User', username, policy_name)
                    if policy_info:
                        inline_policies.append(policy_info)
            except Exception as e:
                utils.log_warning(f"Error getting inline policies for user {username}: {e}")

    group_paginator = iam_client.get_paginator('list_groups')
    for page in group_paginator.paginate():
        for group in page['Groups']:
            processed += 1
            progress = (processed / total_entities) * 100 if total_entities > 0 else 0
            groupname = group['GroupName']
            utils.log_info(f"[{progress:.1f}%] Checking group {processed}/{total_entities}: {groupname}")

            try:
                group_policies = iam_client.list_group_policies(GroupName=groupname)
                for policy_name in group_policies['PolicyNames']:
                    policy_info = process_inline_policy(iam_client, 'Group', groupname, policy_name)
                    if policy_info:
                        inline_policies.append(policy_info)
            except Exception as e:
                utils.log_warning(f"Error getting inline policies for group {groupname}: {e}")

    role_paginator = iam_client.get_paginator('list_roles')
    for page in role_paginator.paginate():
        for role in page['Roles']:
            processed += 1
            progress = (processed / total_entities) * 100 if total_entities > 0 else 0
            rolename = role['RoleName']
            utils.log_info(f"[{progress:.1f}%] Checking role {processed}/{total_entities}: {rolename}")

            try:
                role_policies = iam_client.list_role_policies(RoleName=rolename)
                for policy_name in role_policies['PolicyNames']:
                    policy_info = process_inline_policy(iam_client, 'Role', rolename, policy_name)
                    if policy_info:
                        inline_policies.append(policy_info)
            except Exception as e:
                utils.log_warning(f"Error getting inline policies for role {rolename}: {e}")

    return inline_policies


def process_inline_policy(iam_client, entity_type, entity_name, policy_name):
    """
    Process a single inline policy.

    Args:
        iam_client: The boto3 IAM client
        entity_type: 'User', 'Group', or 'Role'
        entity_name: Name of the entity
        policy_name: Name of the inline policy

    Returns:
        dict: Policy information or None if error
    """
    try:
        if entity_type == 'User':
            response = iam_client.get_user_policy(UserName=entity_name, PolicyName=policy_name)
        elif entity_type == 'Group':
            response = iam_client.get_group_policy(GroupName=entity_name, PolicyName=policy_name)
        elif entity_type == 'Role':
            response = iam_client.get_role_policy(RoleName=entity_name, PolicyName=policy_name)
        else:
            return None

        policy_doc = response['PolicyDocument']
        analysis = analyze_policy_document(policy_doc)

        pseudo_arn = f"inline-policy::{entity_type.lower()}::{entity_name}::{policy_name}"
        attached_users = entity_name if entity_type == 'User' else 'None'
        attached_groups = entity_name if entity_type == 'Group' else 'None'
        attached_roles = entity_name if entity_type == 'Role' else 'None'

        return {
            'Policy Name': f"{policy_name} ({entity_type}: {entity_name})",
            'Policy Type': 'Inline',
            'Policy ARN': pseudo_arn,
            'Policy ID': f"inline-{entity_type.lower()}-{entity_name}-{policy_name}",
            'Attached To Count': 1,
            'Attached Users': attached_users,
            'Attached Groups': attached_groups,
            'Attached Roles': attached_roles,
            'Permission Summary': analysis['permission_summary'],
            'Resource Scope': analysis['resource_scope'],
            'Has Wildcard Actions': analysis['has_wildcard_actions'],
            'Has Wildcard Resources': analysis['has_wildcard_resources'],
            'Statement Count': analysis['statement_count'],
            'Condition Usage': analysis['condition_usage'],
            'Version': 'N/A',
            'Default Version ID': 'N/A',
            'Creation Date': 'Unknown',
            'Last Updated': 'Unknown',
            'Days Since Last Updated': 'Unknown',
            'Path': 'N/A',
            'Description': f"Inline policy attached to {entity_type.lower()}: {entity_name}",
            'Usage Status': 'Used',
            'Risk Level': analysis['risk_level']
        }

    except Exception as e:
        utils.log_warning(f"Error processing inline policy {policy_name} for {entity_type} {entity_name}: {e}")
        return None


# ---------------------------------------------------------------------------
# Export functions (private, prefixed with _)
# ---------------------------------------------------------------------------

def _export_users_to_excel(user_data, account_id, account_name):
    """
    Export IAM user data to Excel file.

    Args:
        user_data: List of user information dictionaries
        account_id: AWS account ID
        account_name: AWS account name

    Returns:
        str: Filename of exported file or None if failed
    """
    if not user_data:
        utils.log_warning("No IAM user data to export.")
        return None

    try:
        import pandas as pd

        df = pd.DataFrame(user_data)

        current_date = datetime.datetime.now().strftime("%m.%d.%Y")
        filename = utils.create_export_filename(account_name, "iam-users", "", current_date)

        summary_data = {
            'Metric': [
                'Total Users',
                'Users with Console Access',
                'Users with MFA Enabled',
                'Users with Access Keys',
                'Users Never Signed In'
            ],
            'Count': [
                len(df),
                len(df[df['Console Access'] == 'Enabled']),
                len(df[df['MFA'] == 'Enabled']),
                len(df[df['Access Key ID'] != 'None']),
                len(df[df['Console Last Sign-in'] == 'Never'])
            ]
        }

        summary_df = pd.DataFrame(summary_data)
        data_frames = {
            'IAM Users': df,
            'Summary': summary_df
        }

        output_path = utils.save_multiple_dataframes_to_excel(data_frames, filename)

        if output_path:
            utils.log_success("AWS IAM user data exported successfully!")
            utils.log_info(f"File location: {output_path}")
            utils.log_info(f"Export contains data for {len(user_data)} IAM users")
            return str(output_path)
        else:
            utils.log_error("Error exporting to Excel. Please check the logs.")
            return None

    except Exception as e:
        utils.log_error("Error exporting to Excel", e)
        return None


def _export_roles_to_excel(role_data, account_id, account_name):
    """
    Export IAM role data to Excel file.

    Args:
        role_data: List of role information dictionaries
        account_id: AWS account ID
        account_name: AWS account name

    Returns:
        str: Filename of exported file or None if failed
    """
    if not role_data:
        utils.log_warning("No IAM role data to export.")
        return None

    try:
        import pandas as pd

        df = pd.DataFrame(role_data)
        df = utils.sanitize_for_export(utils.prepare_dataframe_for_export(df))

        current_date = datetime.datetime.now().strftime("%m.%d.%Y")
        filename = utils.create_export_filename(account_name, "iam-roles", "", current_date)

        summary_data = {
            'Metric': [
                'Total Roles',
                'Service-linked Roles',
                'Cross-account Roles',
                'Standard Roles',
                'Unused Roles (Never Used)',
                'Unused Roles (>90 days)',
                'Roles with Cross-Account Access',
                'Roles with Multiple Policies'
            ],
            'Count': [
                len(df),
                len(df[df['Role Type'] == 'Service-linked']),
                len(df[df['Role Type'] == 'Cross-account']),
                len(df[df['Role Type'] == 'Standard']),
                len(df[df['Last Used'] == 'Never']),
                len(df[(df['Days Since Last Used'] != 'Never') & (df['Days Since Last Used'] != 'Unknown') & (pd.to_numeric(df['Days Since Last Used'], errors='coerce') > 90)]),
                len(df[df['Cross-Account Access'].str.startswith('Yes', na=False)]),
                len(df[df['Permission Policies'].str.contains(',', na=False)])
            ]
        }

        summary_df = pd.DataFrame(summary_data)
        summary_df = utils.sanitize_for_export(utils.prepare_dataframe_for_export(summary_df))

        data_frames = {
            'IAM Roles': df,
            'Summary': summary_df
        }

        output_path = utils.save_multiple_dataframes_to_excel(data_frames, filename)

        if output_path:
            utils.log_success("AWS IAM role data exported successfully!")
            utils.log_info(f"File location: {output_path}")
            utils.log_info(f"Export contains data for {len(role_data)} IAM roles")
            return str(output_path)
        else:
            utils.log_error("Error exporting to Excel. Please check the logs.")
            return None

    except Exception as e:
        utils.log_error("Error exporting to Excel", e)
        return None


def _export_policies_to_excel(managed_policies, inline_policies, account_id, account_name):
    """
    Export IAM policy data to Excel file with risk-level color formatting.

    Args:
        managed_policies: List of managed policy dictionaries
        inline_policies: List of inline policy dictionaries
        account_id: AWS account ID
        account_name: AWS account name

    Returns:
        str: Filename of exported file or None if failed
    """
    if not managed_policies and not inline_policies:
        utils.log_warning("No IAM policy data to export.")
        return None

    try:
        import pandas as pd

        current_date = datetime.datetime.now().strftime("%m.%d.%Y")
        filename = utils.create_export_filename(account_name, "iam-policies", "", current_date)

        data_frames = {}

        if managed_policies:
            managed_df = pd.DataFrame(managed_policies)
            managed_df = utils.sanitize_for_export(utils.prepare_dataframe_for_export(managed_df))
            data_frames['Customer Managed Policies'] = managed_df

        if inline_policies:
            inline_df = pd.DataFrame(inline_policies)
            inline_df = utils.sanitize_for_export(utils.prepare_dataframe_for_export(inline_df))
            data_frames['Inline Policies'] = inline_df

        all_policies = managed_policies + inline_policies
        all_df = pd.DataFrame(all_policies) if all_policies else pd.DataFrame()

        if not all_df.empty:
            total_policies = len(all_df)
            unused_policies = len(all_df[all_df['Usage Status'] == 'Unused'])
            high_risk_policies = len(all_df[all_df['Risk Level'] == 'High'])
            medium_risk_policies = len(all_df[all_df['Risk Level'] == 'Medium'])
            policies_with_wildcards = len(all_df[(all_df['Has Wildcard Actions'] == 'Yes') | (all_df['Has Wildcard Resources'] == 'Yes')])
            policies_without_conditions = len(all_df[all_df['Condition Usage'] == 'No'])

            max_attachments = all_df['Attached To Count'].max() if 'Attached To Count' in all_df.columns else 0
            most_attached = all_df[all_df['Attached To Count'] == max_attachments]['Policy Name'].iloc[0] if max_attachments > 0 else 'None'

            recently_created = 0
            old_policies = 0
            if managed_policies:
                mgd_df = pd.DataFrame(managed_policies)
                for _, policy in mgd_df.iterrows():
                    try:
                        if policy['Days Since Last Updated'] != 'Unknown':
                            days = int(policy['Days Since Last Updated'])
                            if days <= 30:
                                recently_created += 1
                            elif days > 365:
                                old_policies += 1
                    except (ValueError, TypeError):
                        continue

            summary_data = {
                'Metric': [
                    'Total Policies',
                    'Customer Managed Policies',
                    'Inline Policies',
                    'Unused Policies',
                    'High Risk Policies',
                    'Medium Risk Policies',
                    'Policies with Wildcards',
                    'Policies without Conditions',
                    'Most Attached Policy',
                    'Max Attachments',
                    'Recently Created (30 days)',
                    'Due for Review (>365 days)'
                ],
                'Count/Value': [
                    total_policies,
                    len(managed_policies),
                    len(inline_policies),
                    unused_policies,
                    high_risk_policies,
                    medium_risk_policies,
                    policies_with_wildcards,
                    policies_without_conditions,
                    most_attached,
                    max_attachments,
                    recently_created,
                    old_policies
                ]
            }
        else:
            summary_data = {
                'Metric': ['Total Policies'],
                'Count/Value': [0]
            }

        summary_df = pd.DataFrame(summary_data)
        summary_df = utils.sanitize_for_export(utils.prepare_dataframe_for_export(summary_df))
        data_frames['Summary'] = summary_df

        output_path = utils.save_multiple_dataframes_to_excel(data_frames, filename)

        if output_path:
            # Apply openpyxl risk-level color formatting
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import PatternFill

                wb = load_workbook(output_path)

                for sheet_name in ['Customer Managed Policies', 'Inline Policies']:
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        risk_col = None

                        for col in range(1, ws.max_column + 1):
                            if ws.cell(row=1, column=col).value == 'Risk Level':
                                risk_col = col
                                break

                        if risk_col:
                            for row in range(2, ws.max_row + 1):
                                risk_value = ws.cell(row=row, column=risk_col).value
                                if risk_value == 'High':
                                    ws.cell(row=row, column=risk_col).fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
                                elif risk_value == 'Medium':
                                    ws.cell(row=row, column=risk_col).fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
                                elif risk_value == 'Low':
                                    ws.cell(row=row, column=risk_col).fill = PatternFill(start_color='E0FFE0', end_color='E0FFE0', fill_type='solid')

                wb.save(output_path)

            except Exception as e:
                utils.log_warning(f"Could not apply conditional formatting: {e}")

            utils.log_success("AWS IAM policy data exported successfully!")
            utils.log_info(f"File location: {output_path}")
            utils.log_info(f"Export contains {len(managed_policies)} managed and {len(inline_policies)} inline policies")
            return str(output_path)
        else:
            utils.log_error("Error exporting to Excel. Please check the logs.")
            return None

    except Exception as e:
        utils.log_error("Error exporting to Excel", e)
        return None


def _export_comprehensive_to_excel(users_data, roles_data, policies_data, account_id, account_name):
    """
    Export comprehensive IAM data (users, roles, policies) to a single Excel workbook.

    Args:
        users_data: List of user information dictionaries
        roles_data: List of role information dictionaries
        policies_data: List of policy information dictionaries
        account_id: AWS account ID
        account_name: AWS account name

    Returns:
        str: Filename of exported file or None if failed
    """
    if not users_data and not roles_data and not policies_data:
        utils.log_warning("No IAM data to export.")
        return None

    try:
        import pandas as pd

        current_date = datetime.datetime.now().strftime("%m.%d.%Y")
        filename = utils.create_export_filename(account_name, "iam-comprehensive", "", current_date)

        data_frames = {}

        if users_data:
            users_df = pd.DataFrame(users_data)
            users_df = utils.sanitize_for_export(utils.prepare_dataframe_for_export(users_df))
            data_frames['IAM Users'] = users_df

        if roles_data:
            roles_df = pd.DataFrame(roles_data)
            roles_df = utils.sanitize_for_export(utils.prepare_dataframe_for_export(roles_df))
            data_frames['IAM Roles'] = roles_df

        if policies_data:
            policies_df = pd.DataFrame(policies_data)
            policies_df = utils.sanitize_for_export(utils.prepare_dataframe_for_export(policies_df))
            data_frames['IAM Policies'] = policies_df

        summary_data = {
            'Category': [
                'IAM Users',
                'IAM Users - Active Console Access',
                'IAM Users - MFA Enabled',
                'IAM Users - Never Signed In',
                'IAM Users - With Access Keys',
                '',
                'IAM Roles',
                'IAM Roles - Service-linked',
                'IAM Roles - Cross-account',
                'IAM Roles - Standard',
                'IAM Roles - Never Used',
                'IAM Roles - Cross-Account Access',
                '',
                'IAM Policies',
                'IAM Policies - Customer Managed',
                'IAM Policies - Unused',
                'IAM Policies - High Risk',
                'IAM Policies - With Wildcards'
            ],
            'Count': [
                len(users_data),
                len([u for u in users_data if u.get('Console Access') == 'Enabled']) if users_data else 0,
                len([u for u in users_data if u.get('MFA') == 'Enabled']) if users_data else 0,
                len([u for u in users_data if u.get('Console Last Sign-in') == 'Never']) if users_data else 0,
                len([u for u in users_data if u.get('Access Key ID', 'None') != 'None']) if users_data else 0,
                '',
                len(roles_data),
                len([r for r in roles_data if r.get('Role Type') == 'Service-linked']) if roles_data else 0,
                len([r for r in roles_data if r.get('Role Type') == 'Cross-account']) if roles_data else 0,
                len([r for r in roles_data if r.get('Role Type') == 'Standard']) if roles_data else 0,
                len([r for r in roles_data if r.get('Last Used') == 'Never']) if roles_data else 0,
                len([r for r in roles_data if r.get('Cross-Account Access', '').startswith('Yes')]) if roles_data else 0,
                '',
                len(policies_data),
                len([p for p in policies_data if p.get('Policy Type') == 'Customer Managed']) if policies_data else 0,
                len([p for p in policies_data if p.get('Usage Status') == 'Unused']) if policies_data else 0,
                len([p for p in policies_data if p.get('Risk Level') == 'High']) if policies_data else 0,
                len([p for p in policies_data if p.get('Has Wildcard Actions') == 'Yes' or p.get('Has Wildcard Resources') == 'Yes']) if policies_data else 0
            ]
        }

        summary_df = pd.DataFrame(summary_data)
        data_frames['Summary'] = summary_df

        output_path = utils.save_multiple_dataframes_to_excel(data_frames, filename)

        if output_path:
            utils.log_success("AWS comprehensive IAM data exported successfully!")
            utils.log_info(f"File location: {output_path}")
            utils.log_info(f"Export contains {len(users_data)} users, {len(roles_data)} roles, and {len(policies_data)} policies")
            return str(output_path)
        else:
            utils.log_error("Error exporting to Excel. Please check the logs.")
            return None

    except Exception as e:
        utils.log_error("Error exporting to Excel", e)
        return None


# ---------------------------------------------------------------------------
# Run functions
# ---------------------------------------------------------------------------

def _run_users_export(account_id, account_name):
    """Collect IAM users and export to Excel."""
    utils.log_info("Starting IAM user information collection from AWS...")
    user_data = collect_iam_user_information()

    if not user_data:
        utils.log_warning("No IAM user data collected.")
        return

    _export_users_to_excel(user_data, account_id, account_name)


def _run_roles_export(account_id, account_name):
    """Collect IAM roles and export to Excel."""
    utils.log_info("Starting IAM role information collection from AWS...")
    role_data = collect_iam_role_information()

    if not role_data:
        utils.log_warning("No IAM role data collected.")
        return

    _export_roles_to_excel(role_data, account_id, account_name)


def _run_policies_export(account_id, account_name, include_aws_managed=False):
    """Collect IAM policies and export to Excel."""
    utils.log_info("Starting IAM policy information collection from AWS...")

    home_region = utils.get_partition_default_region()
    iam_client = utils.get_boto3_client('iam', region_name=home_region)

    managed_policies = collect_managed_policies(iam_client, include_aws_managed)
    inline_policies = collect_inline_policies(iam_client)

    if not managed_policies and not inline_policies:
        utils.log_warning("No IAM policy data collected.")
        return

    _export_policies_to_excel(managed_policies, inline_policies, account_id, account_name)


def _run_comprehensive_export(account_id, account_name):
    """Collect all IAM resources and export to a single comprehensive workbook."""
    utils.log_info("Starting comprehensive IAM information collection from AWS...")

    utils.log_info("Phase 1: Collecting IAM Users...")
    users_data = collect_iam_user_information()

    utils.log_info("Phase 2: Collecting IAM Roles...")
    roles_data = collect_iam_role_information()

    utils.log_info("Phase 3: Collecting IAM Policies (Customer Managed)...")
    home_region = utils.get_partition_default_region()
    iam_client = utils.get_boto3_client('iam', region_name=home_region)
    policies_data = collect_managed_policies(iam_client, include_aws_managed=False)

    if not users_data and not roles_data and not policies_data:
        utils.log_warning("No IAM data collected.")
        return

    _export_comprehensive_to_excel(users_data, roles_data, policies_data, account_id, account_name)


# ---------------------------------------------------------------------------
# Main — state machine with b/x navigation
# ---------------------------------------------------------------------------

def main():
    """Main function — state machine with b/x navigation and internal IAM menu."""
    try:
        if not utils.ensure_dependencies('pandas', 'openpyxl', 'boto3'):
            return

        utils.setup_logging("iam-export")
        account_id, account_name = utils.print_script_banner("AWS IAM RESOURCES EXPORT")

        step = 1
        choice = None
        include_aws_managed = False

        while True:
            if step == 1:
                result = utils.prompt_menu(
                    "IAM EXPORT OPTIONS",
                    [
                        "IAM Users",
                        "IAM Roles",
                        "IAM Policies",
                        "All IAM Resources (Users + Roles + Policies)",
                    ],
                )
                if result == 'back':
                    sys.exit(10)
                if result == 'exit':
                    sys.exit(11)
                choice = result
                step = 2

            elif step == 2:
                # Policy scope sub-menu only for choice 3 (Policies)
                if choice == 3:
                    result = utils.prompt_menu(
                        "POLICY SCOPE",
                        [
                            "Customer Managed Policies only",
                            "Customer Managed + AWS Managed Policies",
                        ],
                    )
                    if result == 'back':
                        step = 1
                        continue
                    if result == 'exit':
                        sys.exit(11)
                    include_aws_managed = (result == 2)
                    step = 3
                else:
                    step = 3

            elif step == 3:
                choice_labels = {
                    1: "IAM Users",
                    2: "IAM Roles",
                    3: f"IAM Policies ({'Customer + AWS Managed' if include_aws_managed else 'Customer Managed only'})",
                    4: "All IAM Resources (Users + Roles + Policies)",
                }
                msg = f"Ready to export: {choice_labels[choice]}."
                result = utils.prompt_confirmation(msg)
                if result == 'back':
                    step = 2 if choice == 3 else 1
                    continue
                if result == 'exit':
                    sys.exit(11)
                step = 4

            elif step == 4:
                if choice == 1:
                    _run_users_export(account_id, account_name)
                elif choice == 2:
                    _run_roles_export(account_id, account_name)
                elif choice == 3:
                    _run_policies_export(account_id, account_name, include_aws_managed)
                elif choice == 4:
                    _run_comprehensive_export(account_id, account_name)

                print("\nScript execution completed.")
                break

    except KeyboardInterrupt:
        print("\n\nScript interrupted by user. Exiting...")
        sys.exit(0)
    except SystemExit:
        raise
    except Exception as e:
        utils.log_error("Unexpected error occurred", e)
        sys.exit(1)


if __name__ == "__main__":
    main()
