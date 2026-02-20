#!/usr/bin/env python3

"""
===========================
= AWS RESOURCE SCANNER =
===========================

Title: AWS Account Billing Data Export
Date: MAR-04-2025

Description:
Exports AWS billing data for specified time periods (monthly or last 12 months), 
organized by service and associated cost. Handles AWS Cost Explorer
limitations for historical data access and provides alternatives
for accessing older billing data.
"""

import os
import sys
import datetime
import re
from dateutil.relativedelta import relativedelta
from botocore.exceptions import ClientError
from pathlib import Path

# Add path to import utils module
try:
    # Try to import directly (if utils.py is in Python path)
    import utils
except ImportError:
    # If import fails, try to find the module relative to this script
    script_dir = Path(__file__).parent.absolute()
    
    # Check if we're in the scripts directory
    if script_dir.name.lower() == 'scripts':
        # Add the parent directory (StratusScan root) to the path
        sys.path.append(str(script_dir.parent))
    else:
        # Add the current directory to the path
        sys.path.append(str(script_dir))
    
    # Try import again
    try:
        import utils
    except ImportError:
        print("ERROR: Could not import the utils module. Make sure utils.py is in the StratusScan directory.")
        sys.exit(1)

# Setup logging
logger = utils.setup_logging('billing-export')
def print_title():
    """
    Print the script title banner and get account info.
    
    Returns:
        tuple: (account_id, account_name)
    """
    print("====================================================================")
    print("                  AWS RESOURCE SCANNER                              ")
    print("====================================================================")
    print("AWS ACCOUNT BILLING DATA EXPORT TOOL")
    print("====================================================================")
    print("====================================================================")
    
    # Get the current AWS account ID
    try:
        # Create a new STS client to get the current account ID
        sts_client = utils.get_boto3_client('sts')
        # Get account ID from caller identity
        account_id = sts_client.get_caller_identity()['Account']
        # Map the account ID to an account name using utils module
        account_name = utils.get_account_name(account_id, default=account_id)
        
        print(f"Account ID: {account_id}")
        print(f"Account Name: {account_name}")
    except Exception as e:
        print(f"Could not determine account information: {e}")
        account_id = "UNKNOWN"
        account_name = "UNKNOWN-ACCOUNT"
    
    print("====================================================================")
    return account_id, account_name

def validate_date_input(date_input):
    """
    Validate user input for last 12 months or month-year.
    
    Args:
        date_input (str): User input string
        
    Returns:
        tuple: (is_valid, is_year_only, start_date, end_date)
    """
    # Define regex patterns
    month_year_pattern = r'^(0[1-9]|1[0-2])-\d{4}$'  # MM-YYYY
    last_12_pattern = r'^last\s*12$'  # "last 12" with flexible spacing
    
    if re.match(last_12_pattern, date_input.lower()):
        # Last 12 months
        today = datetime.datetime.now()
        end_date = datetime.datetime(today.year, today.month, 1) - datetime.timedelta(days=1)  # Last day of previous month
        start_date = datetime.datetime(end_date.year - 1, end_date.month, 1)  # 12 months before start of previous month
        return True, False, start_date, end_date
    
    elif re.match(month_year_pattern, date_input):
        # Month-Year format (MM-YYYY)
        month, year = date_input.split('-')
        month = int(month)
        year = int(year)
        
        start_date = datetime.datetime(year, month, 1)
        # Calculate the last day of the month
        if month == 12:
            end_date = datetime.datetime(year + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            end_date = datetime.datetime(year, month + 1, 1) - datetime.timedelta(days=1)
        
        return True, False, start_date, end_date
    
    else:
        return False, None, None, None

def check_cost_explorer_data_retention():
    """
    Check if extended data retention is enabled for Cost Explorer.
    
    Returns:
        tuple: (has_extended_retention, max_months)
    """
    try:
        # Create a Cost Explorer client
        ce_client = utils.get_boto3_client('ce')

        # Get Cost Explorer preferences
        response = ce_client.get_preference('COST_EXPLORER')
        
        # Check if extended data retention is enabled
        if 'retentionPeriod' in response:
            retention_period = response['retentionPeriod']
            if retention_period.get('retention') == 'LIFETIME':
                return True, 28  # 14 (standard) + 14 (extended)
            
        # Default retention if not explicitly set
        return False, 14
    except Exception:
        # If we can't determine, assume standard retention
        return False, 14

def validate_date_range(start_date, end_date):
    """
    Validate the date range against AWS Cost Explorer limitations.
    
    Args:
        start_date (datetime): Start date
        end_date (datetime): End date
        
    Returns:
        tuple: (is_valid, message, retention_months)
    """
    today = datetime.datetime.now()
    
    # Cost Explorer data is available the next day
    latest_available_date = today - datetime.timedelta(days=1)
    
    # Check if extended data retention is enabled
    has_extended_retention, retention_months = check_cost_explorer_data_retention()
    
    # Calculate earliest available date based on retention period
    earliest_available_date = today - datetime.timedelta(days=retention_months * 30)
    
    # Check if end date is in the future
    if end_date > latest_available_date:
        end_date_str = end_date.strftime('%Y-%m-%d')
        latest_date_str = latest_available_date.strftime('%Y-%m-%d')
        return False, f"End date ({end_date_str}) is in the future. Latest available data is for {latest_date_str}.", retention_months
    
    # Check if start date is too far in the past
    if start_date < earliest_available_date:
        start_date_str = start_date.strftime('%Y-%m-%d')
        earliest_date_str = earliest_available_date.strftime('%Y-%m-%d')
        retention_text = f"{retention_months} months"
        if has_extended_retention:
            retention_text += " (with extended retention enabled)"
        else:
            retention_text += " (standard retention)"
        
        return False, f"Start date ({start_date_str}) is too far in the past. AWS Cost Explorer only provides data for {retention_text}, available from {earliest_date_str}.", retention_months
    
    return True, "Date range is valid.", retention_months

def get_billing_data(start_date, end_date):
    """
    Get billing data from AWS Cost Explorer API.
    
    Args:
        start_date (datetime): Start date
        end_date (datetime): End date
        
    Returns:
        dict: Billing data organized by month and service
    """
    # Convert dates to string format required by AWS API
    start_date_str = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')
    
    print(f"Fetching billing data from {start_date_str} to {end_date_str}...")
    
    # Create a Cost Explorer client
    ce_client = utils.get_boto3_client('ce')

    try:
        # Use the cost explorer API to get cost and usage data
        response = ce_client.get_cost_and_usage(
            TimePeriod={
                'Start': start_date_str,
                'End': end_date_str
            },
            Granularity='MONTHLY',
            Metrics=['BlendedCost'],
            GroupBy=[
                {
                    'Type': 'DIMENSION',
                    'Key': 'SERVICE'
                }
            ]
        )
        
        # Organize the data by month and service
        billing_data = {}
        
        for result in response['ResultsByTime']:
            # Extract the month from the time period
            period_start = result['TimePeriod']['Start']
            month = datetime.datetime.strptime(period_start, '%Y-%m-%d').strftime('%Y-%m')
            
            # Initialize the month in the billing data if not already present
            if month not in billing_data:
                billing_data[month] = {}
            
            # Process each service and its cost
            for group in result['Groups']:
                service_name = group['Keys'][0]
                cost = float(group['Metrics']['BlendedCost']['Amount'])
                
                # Add the service cost to the month data
                billing_data[month][service_name] = cost
        
        return billing_data
        
    except ClientError as e:
        error_code = e.response.get('Error', {}).get('Code', '')
        error_message = e.response.get('Error', {}).get('Message', str(e))
        
        if error_code == 'ValidationException' and 'historical data' in error_message:
            print("\nError: AWS Cost Explorer cannot access historical data beyond the default 14 months.")
            print("This is a limitation of AWS Cost Explorer. To access older data:")
            print("1. Enable extended data retention in Cost Explorer settings (up to 14 additional months)")
            print("   - Sign in to AWS Management Console")
            print("   - Go to AWS Cost Management > Cost Explorer > Settings")
            print("   - Under 'Data retention', enable 'Extended retention'")
            print("2. For data older than 28 months, set up AWS Cost and Usage Reports (CUR)")
            print("   - Go to AWS Cost Management > Cost & Usage Reports")
            print("   - Set up a report to be delivered to an S3 bucket")
            sys.exit(1)
        else:
            print(f"\nError accessing Cost Explorer: {error_message}")
            sys.exit(1)

def create_excel_report(billing_data, account_name, date_suffix):
    """
    Create an Excel report with monthly billing data.
    
    Args:
        billing_data (dict): Billing data organized by month and service
        account_name (str): Name of AWS account for file naming
        date_suffix (str): Date suffix for filename
        
    Returns:
        str: Path to the created Excel file
    """
    # Import required modules here
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    import pandas as pd
    
    # Create a new workbook
    wb = Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Sort months chronologically
    sorted_months = sorted(billing_data.keys())
    
    # Create a summary sheet
    summary_sheet = wb.create_sheet("Summary")
    summary_sheet['A1'] = 'Month'
    summary_sheet['B1'] = 'Total Cost (USD)'
    
    # Apply header styles to summary sheet
    summary_sheet['A1'].font = header_font
    summary_sheet['B1'].font = header_font
    summary_sheet['A1'].fill = header_fill
    summary_sheet['B1'].fill = header_fill
    
    summary_row = 2
    total_all_months = 0
    
    # Process each month
    for month in sorted_months:
        # Create a sheet for the month
        sheet_name = datetime.datetime.strptime(month, '%Y-%m').strftime('%b %Y')
        ws = wb.create_sheet(sheet_name)
        
        # Create headers
        ws['A1'] = 'Service'
        ws['B1'] = 'Cost (USD)'
        
        # Apply header styles
        ws['A1'].font = header_font
        ws['B1'].font = header_font
        ws['A1'].fill = header_fill
        ws['B1'].fill = header_fill
        
        # Get monthly data
        month_data = billing_data[month]
        
        # Sort services by cost (descending)
        sorted_services = sorted(month_data.items(), key=lambda x: x[1], reverse=True)
        
        # Add data rows
        row = 2
        total_cost = 0
        
        for service, cost in sorted_services:
            ws[f'A{row}'] = service
            ws[f'B{row}'] = cost
            ws[f'B{row}'].number_format = '$#,##0.00'
            total_cost += cost
            row += 1
        
        # Add total row
        row += 1
        ws[f'A{row}'] = 'Total'
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'] = total_cost
        ws[f'B{row}'].font = header_font
        ws[f'B{row}'].number_format = '$#,##0.00'
        
        # Adjust column widths
        for col in range(1, 3):
            column_letter = get_column_letter(col)
            max_length = 0
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add entry to summary sheet
        display_month = datetime.datetime.strptime(month, '%Y-%m').strftime('%B %Y')
        summary_sheet[f'A{summary_row}'] = display_month
        summary_sheet[f'B{summary_row}'] = total_cost
        summary_sheet[f'B{summary_row}'].number_format = '$#,##0.00'
        summary_row += 1
        total_all_months += total_cost
    
    # Add total row to summary
    summary_sheet[f'A{summary_row}'] = 'Total All Months'
    summary_sheet[f'B{summary_row}'] = total_all_months
    summary_sheet[f'A{summary_row}'].font = header_font
    summary_sheet[f'B{summary_row}'].font = header_font
    summary_sheet[f'B{summary_row}'].number_format = '$#,##0.00'
    
    # Adjust summary sheet column widths
    for col in range(1, 3):
        column_letter = get_column_letter(col)
        max_length = 0
        for cell in summary_sheet[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        summary_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Generate filename using utils
    filename = utils.create_export_filename(
        account_name, 
        "billing", 
        date_suffix, 
        datetime.datetime.now().strftime("%m.%d.%Y")
    )
    
    # Get the full output path
    output_path = utils.get_output_filepath(filename)
    
    # Ensure the output directory exists
    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)
    
    # Save the workbook
    wb.save(output_path)
    print(f"Excel report saved as: {output_path}")
    return output_path

def main():
    """
    Main function to run the script.
    """
    try:
        # Check partition availability
        partition = utils.detect_partition()
        if not utils.is_service_available_in_partition("ce", partition):
            utils.log_warning("Cost Explorer (billing) is not available in AWS GovCloud. Skipping.")
            sys.exit(0)

        # Print title and get account info
        account_id, account_name = print_title()
        
        # Check dependencies
        if not utils.ensure_dependencies('pandas', 'openpyxl', 'python-dateutil'):
            sys.exit(1)
        
        # Get user input for date range
        while True:
            date_input = input("\nWould you like the last 12 months (type \"last 12\") or a specific month (ex. \"01-2025\")? ")
            
            is_valid, is_year_only, start_date, end_date = validate_date_input(date_input)
            
            if is_valid:
                # Validate date range against AWS limitations
                date_valid, message, _ = validate_date_range(start_date, end_date)
                if date_valid:
                    break
                else:
                    print(f"Error: {message}")
                    print("Please try again with a more recent date range.")
            else:
                print("Invalid input format. Please enter either \"last 12\" or a month in format \"MM-YYYY\" (e.g., \"01-2025\").")
        
        # Get billing data
        billing_data = get_billing_data(start_date, end_date)
        
        if not billing_data:
            print("\nNo billing data found for the specified period.")
            sys.exit(0)
        
        # Determine output file name suffix
        if date_input.lower().startswith('last'):
            date_suffix = "last-12-months"
        else:
            date_suffix = start_date.strftime('%m-%Y')
        
        # Create Excel report
        output_file = create_excel_report(billing_data, account_name, date_suffix)
        
        print("\nBilling data export completed successfully.")
        print(f"File saved to: {output_file}")
        
    except KeyboardInterrupt:
        print("\nOperation cancelled by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\nAn error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
